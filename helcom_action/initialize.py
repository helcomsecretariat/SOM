# -*- coding: utf-8 -*-
"""
Created on Tue Aug  4 16:13:20 2020

@author: E1005457
"""
#These libraries need to be installed before they can be imported
#There may be some that are not actually needed
from openpyxl import Workbook
from openpyxl import load_workbook
import math as mt
import numpy as np
#import scipy
import pandas as pd
#import statistics as st
from ownFunctions import readData
from ownFunctions import deflambda
from ownFunctions import caus_impact
from ownFunctions import areas

#import plotly.graph_objects as go
#import plotly as py
#import random
from scipy.stats import beta
from scipy import stats
import dill
import seaborn as sns
import matplotlib.pyplot as plt
#import matplotlib.ticker as tkr

#check that the path indicates to right folder, in this case SOM-excels
#General input data
wb = load_workbook('SOM-excels\generalInput.xlsx',data_only=True)
#Measure effects from expert survey
wb1 = load_workbook('SOM-excels\measureEffInput.xlsx',data_only=True)
#Activity pressure contributions
wb3 = load_workbook('SOM-excels\\actPresNew.xlsx',data_only=True)#wb3 = load_workbook('SOM-excels\\actPresInput.xlsx',data_only=True)
#Litter results, calculated separately
wb4 = load_workbook('SOM-excels\Litter_to_state.xlsx',data_only=True)
#Nutrient reduction results, calculated  separately
wb5 = load_workbook('SOM-excels\inputNutrients.xlsx',data_only=True)
#Pressure-State input from expert survey
wb2 = load_workbook('SOM-excels\pressStateInput.xlsx',data_only=True)
#Literature input to replace expert data
wb7 = load_workbook('SOM-excels\lit_ME_Liisa.xlsx',data_only=True)
#Effectiveness for new measures
wb8 = load_workbook('SOM-excels\\new_measures_new.xlsx',data_only=True)
